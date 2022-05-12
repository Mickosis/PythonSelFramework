import openpyxl

book = openpyxl.load_workbook("/xxxx/test_data.xlsx")
sheet = book.active
dict_values = {}
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
# sheet.cell(row=4, column=2).value = "Test"
#
# print(sheet.cell(row=4, column=2).value)
#
# print(sheet.max_row)
#
# print(sheet.max_column)
#
# print(sheet['A1'].value)

for i in range(2, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == 2:
        for j in range(2, sheet.max_column + 1):
            dict_values[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
