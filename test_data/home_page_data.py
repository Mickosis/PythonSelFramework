import openpyxl


class HomePageData:
    test_home_page_data = [{"firstname": "XXXX", "lastname": "XXXX", "gender": "Male"},
                           {"firstname": "XXXX", "lastname": "XXXX", "gender": "Female"}]

    @staticmethod
    def get_excel_data(test_case_name):

        book = openpyxl.load_workbook("/xxxx/test_data.xlsx")
        sheet = book.active
        dict_values = {}

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict_values[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict_values]
